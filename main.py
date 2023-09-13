from openpyxl import load_workbook
import base64
import os
from PIL import Image
import  json
import requests


workbook = load_workbook("hello.xlsx")
sheet = workbook.active
API_ENDPOINT = ""
row01 = sheet[2]
images_dir = ''
image_list = []

def simplifyname(name:str):
    colors = ["GRAY","BLANCHE","BLANC","BLACK","WHITE","GRIS","GRISE","NOIR"]
    name = name.upper()
    for color in colors:
       name = name.replace("/"+color,"")
       name = name.replace("/ " + color, "")
    return name.lower().title()


# Cinvert categories to ids
def convert_id(string):
    switch_dict = {
        "climatiseur": "08da8f28-426a-4236-8502-e0d2f415597c",
        "lavage": "08da8f28-4dcf-46da-8a1b-72fdf287930b",
        "cuisson": "08da8f28-5b86-400b-8ae1-e585ab1e4994",
        "chauffage": "08da8f28-6937-4df0-87ca-bdffb9fcc4a5",
        "tv": "3fa85f64-5717-4562-b3fc-2c963f66afa6",
        "laptop": "08daa163-dcdb-4e51-8c7a-def79d939bae",
        "all in one": "08daa164-7c33-47bc-81a8-bb8eec475b3e",
        "froid": "08da8f28-743b-4a8d-8bf1-dc65d4bf4083",
        "pack": "08db64e7-2310-4cb1-8759-6d74b10a3900",
        "pem": "08da8f28-857a-419d-8d42-325aaab69515",
    }

    return switch_dict.get(string.lower(), 1)
def GetColor(name : str):
    colors = ["GRAY","BLANCHE","BLANC","BLACK","WHITE","GRIS","GRISE","NOIR"]
    switch_dict = {
        "GRAY": "#808080",
        "BLANCHE": "#fff",
        "BLANC": "#fff",
        "BLACK": "#000",
        "WHITE": "#fff",
        "GRIS": "#808080",
        "GRISE": "#808080",
        "NOIR": "#000",
    }
    name = name.upper()
    for color in colors:
        if name.find(color) != -1:
             return switch_dict.get(color, 1)
    return "#000"



# Images class
class Variante :
    def __init__(self,name,namear,color,reference,visible,price):
        self.name = name
        self.namear = namear
        self.color = color
        self.reference = reference
        self.visible = visible
        self.price = price


class ImageDto :
    def __init__(self,base64,size,height,width):
        self.base64 = base64
        self.size = size
        self.height = height
        self.width = width


# products class
class Product:
    def __init__(self, name, description,namear, priority, populaire, Visible, CategoryId, brandId,images,variantes):
        self.name = name
        self.description = description
        self.namear = namear
        self.priority = priority
        self.populaire = populaire
        self.visible = Visible
        self.categoryId = convert_id(CategoryId)
        self.brandId = "3fa85f64-5717-4562-b3fc-2c963f66afa6"
        self.images = images
        self.variantes = variantes

ls = []

# read and covert images to base64

def readImages(articleRef):
    for filename in os.listdir(os.path.join(images_dir,articleRef)):
        if filename.endswith(('.jpg', '.jpeg', '.png', '.gif')):
            # Build the full path to the image
            image_path = os.path.join(images_dir, filename)
            # Open the image using Pillow
            with Image.open(images_dir+"/"+articleRef+"/"+filename) as img:
                print(images_dir+"/"+articleRef+"/"+filename)
                b64 = base64.b64encode(img.tobytes()).decode()

                imagedto = ImageDto(filename+","+b64, img.size[0]*img.size[1], img.height, img.width)
                image_list.append(imagedto)
    print(image_list[0].base64)
    return  image_list
# print(row01.values)



def getvariantes(name):
    variantes = []
    i = 1
    for value in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=7, values_only=True):
        i = i+1
        # print(simplifyname(name))
        if value[0] == None:
            break

        # print("product name = " ,simplifyname(sheet[i][3].value))

        if simplifyname(value[3]) == simplifyname(name) :
            variante = Variante(value[3],value[5],GetColor(value[3]),value[2],True,value[4])
            variantes.append(variante)
            # print("---------------------------------")
            # print(variante['name'])
            # # delete the current row
            # print("=======================")
            # print(sheet[i][3].value)
            # # sheet.delete_rows(i)
            # print( "i=", i)
            # print("variantes = ", len(variantes))
            # # print("sheet len = ", sheet. )
            # print("---------------------------------")


    return variantes

for value in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=6, values_only=True):
    if value[0] == None:
        break
    list = []
    for product in ls:
        list.append(product.name)
    if(simplifyname(value[3]) in list):
        continue
    variantes = getvariantes(value[3])
    images = readImages(value[2])
    product = Product(simplifyname(value[3]),simplifyname(value[3]),simplifyname(value[3]),1,True,True,value[1],value[0],[img.__dict__ for img in images],[v.__dict__ for v in variantes])
    ls.append(product)


print("products =" ,len(ls))
output_json_file = 'output.json'
serialized_data = json.dumps([product.__dict__ for product in ls])

with open(output_json_file, 'w', encoding='utf-8') as json_file:
    json.dump(serialized_data, json_file, ensure_ascii=False, indent=4)
headers = {'Content-Type': 'application/json'}
print(ls[0].__dict__)





r = requests.post(url=API_ENDPOINT, data = json.dumps(ls[0].__dict__), headers=headers)
print(ls[0].images[0].base64)
